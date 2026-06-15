"""CSV/XLSX auto-profiler for messy warehouse datasets."""

from __future__ import annotations

import csv
import re
from collections import Counter
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any


ISSUE_TYPES = (
    "TYPE_CHAOS",
    "UNIT_INCONSISTENCY",
    "CATEGORICAL_VARIANTS",
    "DUPLICATES",
    "HIGH_NULL_RATE",
    "FORMAT_INCONSISTENCY",
)

_KG_PATTERN = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*(kg|kilogram|kilograms)?\s*$", re.I)
_DATE_PATTERNS = (
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    re.compile(r"^\d{2}/\d{2}/\d{4}$"),
    re.compile(r"^[A-Za-z]{3}\s+\d{1,2}\s+\d{4}$"),
)


@dataclass(slots=True)
class ColumnSchema:
    name: str
    inferred_type: str
    null_pct: float
    unique_pct: float


@dataclass(slots=True)
class DetectedIssue:
    col: str
    issue_type: str
    examples: list[str]
    row_count: int


@dataclass(slots=True)
class ProfileReport:
    schema: list[ColumnSchema]
    detected_issues: list[DetectedIssue] = field(default_factory=list)
    candidate_pk: list[str] = field(default_factory=list)
    row_count: int = 0

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_count": self.row_count,
            "schema": [asdict(s) for s in self.schema],
            "detected_issues": [asdict(i) for i in self.detected_issues],
            "candidate_pk": self.candidate_pk,
        }


def _read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = list(reader.fieldnames or [])
            return headers, [dict(r) for r in reader]
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) for h in next(rows_iter)]
        rows = []
        for row in rows_iter:
            rows.append({headers[i]: "" if v is None else str(v) for i, v in enumerate(row)})
        return headers, rows
    raise ValueError(f"Unsupported file type: {suffix}")


def _infer_type(values: list[str]) -> str:
    non_empty = [v for v in values if v.strip()]
    if not non_empty:
        return "unknown"
    numeric = sum(1 for v in non_empty if _KG_PATTERN.match(v.strip()) or v.strip().replace(".", "", 1).isdigit())
    if numeric / len(non_empty) > 0.7:
        return "numeric_or_unit"
    date_hits = sum(
        1
        for v in non_empty
        if any(p.match(v.strip()) for p in _DATE_PATTERNS)
    )
    if date_hits / len(non_empty) > 0.5:
        return "date"
    return "string"


def profile_dataset(path: Path | str) -> ProfileReport:
    path = Path(path)
    headers, rows = _read_rows(path)
    report = ProfileReport(schema=[], row_count=len(rows))
    col_values: dict[str, list[str]] = {h: [] for h in headers}

    for row in rows:
        for h in headers:
            col_values[h].append(str(row.get(h, "") or ""))

    for h in headers:
        vals = col_values[h]
        non_empty = [v for v in vals if v.strip()]
        null_pct = round(100.0 * (len(vals) - len(non_empty)) / max(len(vals), 1), 2)
        unique_pct = round(100.0 * len(set(non_empty)) / max(len(non_empty), 1), 2)
        report.schema.append(
            ColumnSchema(
                name=h,
                inferred_type=_infer_type(vals),
                null_pct=null_pct,
                unique_pct=unique_pct,
            )
        )

    # Issue detection
    for h in headers:
        vals = col_values[h]
        non_empty = [v for v in vals if v.strip()]
        null_pct = 100.0 * (len(vals) - len(non_empty)) / max(len(vals), 1)
        if null_pct >= 10:
            report.detected_issues.append(
                DetectedIssue(h, "HIGH_NULL_RATE", non_empty[:3], int(len(vals) - len(non_empty)))
            )

    # SKU format chaos
    if "sku" in col_values:
        skus = [v.strip() for v in col_values["sku"] if v.strip()]
        normalized = {re.sub(r"[\s-]", "", s.upper()) for s in skus}
        if len(normalized) < len(skus) * 0.85:
            examples = list({s for s in skus})[:5]
            report.detected_issues.append(
                DetectedIssue("sku", "FORMAT_INCONSISTENCY", examples, len(skus) - len(normalized))
            )

    # Unit inconsistency
    if "quantity_kg" in col_values:
        units = Counter()
        for v in col_values["quantity_kg"]:
            if not v.strip():
                continue
            m = _KG_PATTERN.match(v.strip())
            if m and m.group(2):
                units[m.group(2).lower()] += 1
            elif m:
                units["bare_number"] += 1
            else:
                units["other"] += 1
        if len(units) > 1:
            report.detected_issues.append(
                DetectedIssue(
                    "quantity_kg",
                    "UNIT_INCONSISTENCY",
                    [f"{k}:{v}" for k, v in units.most_common(3)],
                    sum(units.values()),
                )
            )

    # Location categorical variants
    loc_field = "location_id" if "location_id" in col_values else "location"
    if loc_field in col_values:
        locs = [v.strip() for v in col_values[loc_field] if v.strip()]
        buckets: dict[str, set[str]] = {"WH-A": set(), "WH-B": set()}
        for loc in locs:
            key = re.sub(r"[\s_]", "", loc.lower())
            if key in {"wh-a", "wha", "warehousea"} or ("warehouse" in key and "a" in key):
                buckets["WH-A"].add(loc)
            elif key in {"wh-b", "whb", "warehouseb"} or ("warehouse" in key and "b" in key):
                buckets["WH-B"].add(loc)
            elif loc.upper().startswith("WH-A") or loc.upper() == "LOC-001":
                buckets["WH-A"].add(loc)
            elif loc.upper().startswith("WH-B") or loc.upper() == "LOC-002":
                buckets["WH-B"].add(loc)
        for variants in buckets.values():
            if len(variants) > 1:
                report.detected_issues.append(
                    DetectedIssue(
                        loc_field,
                        "CATEGORICAL_VARIANTS",
                        sorted(variants)[:5],
                        len(variants),
                    )
                )

    # Date format inconsistency
    date_col = "last_restocked" if "last_restocked" in col_values else "last_updated"
    if date_col in col_values:
        formats = Counter()
        for v in col_values[date_col]:
            if not v.strip():
                continue
            if _DATE_PATTERNS[0].match(v.strip()):
                formats["iso"] += 1
            elif _DATE_PATTERNS[1].match(v.strip()):
                formats["dmy"] += 1
            elif _DATE_PATTERNS[2].match(v.strip()):
                formats["text"] += 1
            else:
                formats["other"] += 1
        if len(formats) > 1:
            report.detected_issues.append(
                DetectedIssue(
                    date_col,
                    "FORMAT_INCONSISTENCY",
                    [f"{k}:{v}" for k, v in formats.most_common(3)],
                    sum(formats.values()),
                )
            )

    # Duplicate rows (fuzzy on sku+location)
    loc_col = "location_id" if "location_id" in col_values else "location"
    if "sku" in col_values and loc_col in col_values:
        seen: Counter[str] = Counter()
        for row in rows:
            sku = re.sub(r"[\s-]", "", str(row.get("sku", "")).upper())
            loc = re.sub(r"[\s_]", "", str(row.get(loc_col, "")).lower())
            seen[f"{sku}|{loc}"] += 1
        dup_count = sum(c - 1 for c in seen.values() if c > 1)
        if dup_count:
            report.detected_issues.append(
                DetectedIssue(f"sku+{loc_col}", "DUPLICATES", list(seen.keys())[:3], dup_count)
            )

    reorder_col = "reorder_level_kg" if "reorder_level_kg" in col_values else "reorder_level"
    if reorder_col in col_values:
        types = Counter()
        for v in col_values[reorder_col]:
            if not v.strip():
                types["null"] += 1
            elif v.strip().upper() == "N/A":
                types["na_string"] += 1
            elif v.strip().isdigit():
                types["int"] += 1
            else:
                types["other"] += 1
        if len([t for t in types if t not in {"null"}]) > 1:
            report.detected_issues.append(
                DetectedIssue(
                    reorder_col,
                    "TYPE_CHAOS",
                    [f"{k}:{v}" for k, v in types.most_common(4)],
                    sum(types.values()),
                )
            )

    # Candidate PK
    for s in report.schema:
        if s.unique_pct >= 95 and s.null_pct < 5:
            report.candidate_pk.append(s.name)

    return report
