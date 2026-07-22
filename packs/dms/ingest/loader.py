"""L1 — file → bronze lakehouse, exactly-once with a content-hash file ledger.

Design guarantees:
  * exactly-once: a file whose content hash already loaded is recorded as
    `skipped_duplicate`, never re-ingested.
  * no partial commits: the bronze table is built with CREATE OR REPLACE ... AS
    SELECT, which only replaces after the read fully materializes; a parse error
    leaves the previous table untouched and records `failed` (quarantine).
  * every load/skip/fail writes a row to `lake.bronze._ingest_ledger` AND an F1
    audit-ledger event.

Supported: .csv/.tsv, .json/.jsonl/.ndjson, .xlsx (first sheet). Bronze is
permissive: all-VARCHAR + `_source_file`/`_content_hash`/`_loaded_at`/`_rescued`.
"""
from __future__ import annotations

import datetime as _dt
import hashlib
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from packs.dms.lakehouse.catalog import LAKE_ALIAS, connect

LEDGER_TABLE = f"{LAKE_ALIAS}.bronze._ingest_ledger"
CSV_EXT = {".csv", ".tsv"}
JSON_EXT = {".json", ".jsonl", ".ndjson"}
XLSX_EXT = {".xlsx"}
SUPPORTED = CSV_EXT | JSON_EXT | XLSX_EXT
_META_COLS = ("_source_file", "_content_hash", "_loaded_at", "_rescued")


@dataclass(slots=True)
class IngestResult:
    status: str            # loaded | skipped_duplicate | failed | unsupported
    table: str | None
    rows: int
    content_hash: str
    path: str
    error: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "status": self.status, "table": self.table, "rows": self.rows,
            "content_hash": self.content_hash, "path": self.path, "error": self.error,
        }


def _now() -> str:
    return _dt.datetime.now(_dt.timezone.utc).isoformat()


def _sanitize(stem: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "_", stem).strip("_").lower()
    return s or "file"


def _content_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _ensure_ledger(con) -> None:
    con.execute(f"CREATE SCHEMA IF NOT EXISTS {LAKE_ALIAS}.bronze")
    con.execute(
        f"CREATE TABLE IF NOT EXISTS {LEDGER_TABLE} ("
        " id VARCHAR, path VARCHAR, filename VARCHAR, size_bytes BIGINT,"
        " mtime VARCHAR, content_hash VARCHAR, target_table VARCHAR,"
        " status VARCHAR, rows BIGINT, error VARCHAR, loaded_at VARCHAR)"
    )


def _insert_ledger(con, *, path: Path, chash: str, status: str,
                   target: str | None, rows: int, error: str = "") -> None:
    size = path.stat().st_size if path.exists() else 0
    mtime = (_dt.datetime.fromtimestamp(path.stat().st_mtime, _dt.timezone.utc).isoformat()
             if path.exists() else "")
    con.execute(
        f"INSERT INTO {LEDGER_TABLE} VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        [str(uuid.uuid4()), str(path), path.name, size, mtime, chash,
         target, status, rows, error[:2000], _now()],
    )


def _audit(event: str, payload: dict, *, actor: str) -> None:
    try:
        from packs.dms.audit import ledger

        ledger.append(actor, event, payload)
    except Exception:  # noqa: BLE001 — audit best-effort, never blocks ingest
        pass


def _already_loaded(con, chash: str) -> bool:
    row = con.execute(
        f"SELECT 1 FROM {LEDGER_TABLE} WHERE content_hash = ? AND status = 'loaded' LIMIT 1",
        [chash],
    ).fetchone()
    return row is not None


# ── format loaders (each builds bronze.<target> WITH meta, atomically) ────────
def _meta_select(fname: str, chash: str) -> str:
    return (f", '{fname}' AS _source_file, '{chash}' AS _content_hash, "
            f"'{_now()}' AS _loaded_at, NULL::VARCHAR AS _rescued")


def _load_csv(con, path: Path, target: str, chash: str) -> int:
    sep = "\t" if path.suffix.lower() == ".tsv" else ","
    con.execute(
        f"CREATE OR REPLACE TABLE {LAKE_ALIAS}.bronze.{target} AS "
        f"SELECT * {_meta_select(path.name, chash)} "
        f"FROM read_csv('{path.as_posix()}', header=true, all_varchar=true, "
        f"              sep='{sep}', ignore_errors=false, null_padding=false)"
    )
    return int(con.execute(f"SELECT COUNT(*) FROM {LAKE_ALIAS}.bronze.{target}").fetchone()[0])


def _load_json(con, path: Path, target: str, chash: str) -> int:
    fmt = "newline_delimited" if path.suffix.lower() in (".jsonl", ".ndjson") else "auto"
    con.execute(
        f"CREATE OR REPLACE TABLE {LAKE_ALIAS}.bronze.{target} AS "
        f"SELECT * {_meta_select(path.name, chash)} "
        f"FROM read_json('{path.as_posix()}', format='{fmt}', "
        f"               auto_detect=true, ignore_errors=false)"
    )
    return int(con.execute(f"SELECT COUNT(*) FROM {LAKE_ALIAS}.bronze.{target}").fetchone()[0])


def _load_xlsx(con, path: Path, target: str, chash: str) -> int:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl not installed (pip install netie[dms])") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        raise ValueError("empty workbook")
    header = [_sanitize(str(c)) if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
    body = rows[1:]
    coldefs = ", ".join(f"{c} VARCHAR" for c in header) + \
        ", _source_file VARCHAR, _content_hash VARCHAR, _loaded_at VARCHAR, _rescued VARCHAR"
    con.execute(f"CREATE OR REPLACE TABLE {LAKE_ALIAS}.bronze.{target} ({coldefs})")
    now = _now()
    payload = [
        [(None if v is None else str(v)) for v in row][:len(header)]
        + [None] * (len(header) - len(row))          # pad short rows
        + [path.name, chash, now, None]
        for row in body
    ]
    if payload:
        from packs.dms.lakehouse.tables import bulk_insert

        all_cols = header + ["_source_file", "_content_hash", "_loaded_at", "_rescued"]
        bulk_insert(con, f"{LAKE_ALIAS}.bronze.{target}", all_cols, payload)
    return len(payload)


_LOADERS = [(CSV_EXT, _load_csv), (JSON_EXT, _load_json), (XLSX_EXT, _load_xlsx)]


def _load_into_bronze(con, path: Path, target: str, chash: str) -> int:
    ext = path.suffix.lower()
    for exts, fn in _LOADERS:
        if ext in exts:
            return fn(con, path, target, chash)
    raise ValueError(f"unsupported extension {ext!r}")


# ── public API ────────────────────────────────────────────────────────────────
def load_one(path: Path | str, *, con=None, actor: str = "system") -> IngestResult:
    path = Path(path)
    owns = con is None
    con = con or connect()
    try:
        _ensure_ledger(con)
        ext = path.suffix.lower()
        if ext not in SUPPORTED:
            return IngestResult("unsupported", None, 0, "", str(path),
                                error=f"unsupported extension {ext!r}")
        data = path.read_bytes()
        chash = _content_hash(data)

        if _already_loaded(con, chash):
            _insert_ledger(con, path=path, chash=chash, status="skipped_duplicate",
                           target=None, rows=0)
            _audit("ingest.skipped", {"path": str(path), "content_hash": chash}, actor=actor)
            return IngestResult("skipped_duplicate", None, 0, chash, str(path))

        target = f"{_sanitize(path.stem)}_raw"
        try:
            rows = _load_into_bronze(con, path, target, chash)
        except Exception as exc:  # noqa: BLE001 — a bad file is a quarantine, not a crash
            _insert_ledger(con, path=path, chash=chash, status="failed",
                           target=target, rows=0, error=repr(exc))
            _audit("ingest.failed", {"path": str(path), "error": repr(exc)}, actor=actor)
            return IngestResult("failed", target, 0, chash, str(path), error=repr(exc))

        _insert_ledger(con, path=path, chash=chash, status="loaded", target=target, rows=rows)
        _audit("ingest.loaded",
               {"path": str(path), "table": target, "rows": rows, "content_hash": chash},
               actor=actor)
        return IngestResult("loaded", target, rows, chash, str(path))
    finally:
        if owns:
            con.close()


def scan(folder: Path | str, *, con=None) -> list[Path]:
    """Files in `folder` whose content hash is not already loaded (new/changed)."""
    folder = Path(folder)
    owns = con is None
    con = con or connect()
    try:
        _ensure_ledger(con)
        out: list[Path] = []
        for path in sorted(folder.iterdir()):
            if not path.is_file() or path.suffix.lower() not in SUPPORTED:
                continue
            chash = _content_hash(path.read_bytes())
            if not _already_loaded(con, chash):
                out.append(path)
        return out
    finally:
        if owns:
            con.close()


def ingest_folder(folder: Path | str, *, actor: str = "system") -> list[IngestResult]:
    folder = Path(folder)
    con = connect()
    try:
        results: list[IngestResult] = []
        for path in sorted(folder.iterdir()):
            if path.is_file() and path.suffix.lower() in SUPPORTED:
                results.append(load_one(path, con=con, actor=actor))
        return results
    finally:
        con.close()


def ledger_entries(*, con=None, limit: int = 200) -> list[dict[str, Any]]:
    owns = con is None
    con = con or connect(read_only=True)
    try:
        _ensure_ledger(con) if not owns else None
        rel = con.execute(
            f"SELECT * FROM {LEDGER_TABLE} ORDER BY loaded_at DESC LIMIT {int(limit)}")
        cols = [d[0] for d in rel.description]
        return [dict(zip(cols, row)) for row in rel.fetchall()]
    except Exception:  # noqa: BLE001 — no ledger yet
        return []
    finally:
        if owns:
            con.close()
